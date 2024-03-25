#pylint:disable=C0301
"""
Module for automating orders on Saucedemo website.

Contains the `OrderAutomation` class to handle orders.

Example:
    automation = OrderAutomation("user_credentials.xlsx")
    automation.setup_driver()
    automation.load_excel()
    automation.place_orders()
    automation.save_excel()
    automation.close_driver()
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import openpyxl


class OrderAutomation:
    """
    Class for automating the process of placing orders on the Saucedemo website.

    Attributes:
        excel_file (str): The file path of the Excel file containing the order details.
        driver (webdriver): A WebDriver instance for interacting with the Saucedemo website.
        wait (WebDriverWait): An explicit wait object for handling asynchronous web page loading.
        wb (openpyxl.Workbook): A Workbook object for reading and writing the Excel file.

    Methods:
        setup_driver(self):
            Initializes the WebDriver instance.

        load_excel(self):
            Loads the Excel file into the Workbook object.

        login(self, username, password):
            Logs in to the Saucedemo website using the provided username and password.

        place_orders(self):
            Automates the process of placing orders on the Saucedemo website.

        save_excel(self):
            Saves the changes made to the Excel file.

        close_driver(self):
            Closes the WebDriver instance.

    """
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.driver = None
        self.wait = None
        self.wb = None

    def setup_driver(self):
        """
        Initializes the WebDriver instance.
        """
        self.driver = webdriver.Chrome()
        self.wait = WebDriverWait(self.driver, 5)

    def load_excel(self):
        """
        Loads the Excel file into the Workbook object.
        """
        self.wb = openpyxl.load_workbook(self.excel_file)

    def login(self, username, password):
        """
        Logs in to the Saucedemo website using the provided username and password.
        """
        self.driver.get("https://www.saucedemo.com/v1/")
        username_field = self.driver.find_element(By.ID, "user-name")
        password_field = self.driver.find_element(By.ID, "password")
        login_button = self.driver.find_element(By.ID, "login-button")
        username_field.send_keys(username)
        password_field.send_keys(password)
        login_button.click()

    def place_orders(self): #pylint:disable=R0914
        """
        Automates the process of placing orders on the Saucedemo website.
        """
        sheet_order_details = self.wb["Order Details"]
        for row in sheet_order_details.iter_rows(min_row=2, values_only=True):
            order_id, user_id, product_name, quantity, total_price, order_status = row #pylint:disable=W0612
            self.login(user_id, "secret_sauce")
            self.wait.until(EC.url_contains("/inventory.html"))
            product_link = self.driver.find_element(By.XPATH, f"//div[normalize-space()='{product_name}']")
            product_link.click()
            add_to_cart_button = self.driver.find_element(By.XPATH, "//button[text()='ADD TO CART']")
            add_to_cart_button.click()
            shoppingcart_button = self.driver.find_element(By.CLASS_NAME, "shopping_cart_link")
            shoppingcart_button.click()
            checkout_button = self.driver.find_element(By.XPATH, "//a[normalize-space()='CHECKOUT']")
            checkout_button.click()
            self.driver.find_element(By.ID, "first-name").send_keys("Anni")
            self.driver.find_element(By.ID, "last-name").send_keys("Sharma")
            self.driver.find_element(By.ID, "postal-code").send_keys("808080")
            continue_button = self.driver.find_element(By.XPATH, "//input[@value='CONTINUE']")
            continue_button.click()

            try:
                item_total_element = self.driver.find_element(By.CSS_SELECTOR, "div.summary_subtotal_label")
                item_total = item_total_element.text
                item_total_value = float(item_total.split("$")[1])
                total_price_float = float(total_price)
                if abs(item_total_value - total_price_float) < 0.01:
                    order_status = "Success"
                else:
                    order_status = "Failure"
            except NoSuchElementException:
                order_status = "Failure"

            try:
                finish_button = self.driver.find_element(By.XPATH, "//a[normalize-space()='FINISH']")
                finish_button.click()
            except NoSuchElementException:
                print("Finish button not found.")

            try:
                success_message = self.wait.until(
                    EC.visibility_of_element_located((By.XPATH, "//h2[text()='THANK YOU FOR YOUR ORDER']")))
                if success_message.is_displayed():
                    order_status = "Success"
            except TimeoutException:
                order_status = "Failure"

            sheet_order_details.cell(row=int(order_id) + 1, column=6, value=order_status)
            self.driver.get("https://www.saucedemo.com/v1/")

    def save_excel(self):
        """
        Saves the changes made to the Excel file.
        """
        self.wb.save(self.excel_file)

    def close_driver(self):
        """
        Closes the WebDriver instance.
        """
        if self.driver:
            self.driver.quit()


def main():
    """
    Sets up the OrderAutomation instance,
        initializes the webdriver,
        loads the Excel file, places orders,
        saves changes, and closes the webdriver.
    """
    excel_file = "user_credentials.xlsx"
    automation = OrderAutomation(excel_file)
    automation.setup_driver()
    automation.load_excel()
    try:
        automation.place_orders()
    finally:
        automation.save_excel()
        automation.close_driver()


if __name__ == "__main__":
    main()
