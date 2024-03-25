"""
This module contains a function to log in to Saucedemo with
 a list of user credentials and records the login status in an Excel file.
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import openpyxl

def login_with_users(credentials_file, login_sheet_name):
    """Logs in to Saucedemo with a list of user credentials
      and records the login status in an Excel file.
    """
    # Initialize Chrome webdriver
    driver = webdriver.Chrome()
    driver.get("https://www.saucedemo.com/v1/")

    # Wait for the page to load
    driver.implicitly_wait(5)  # Adjust the waiting time as necessary

    # Load user credentials from the Excel file
    wb = openpyxl.load_workbook(credentials_file)
    sheet_credentials = wb["User credentials"]

    # Create a new sheet for login messages
    sheet_login = wb.create_sheet(login_sheet_name)
    sheet_login['A1'] = "User ID"
    sheet_login['B1'] = "Login Message"

    # Loop through each user credential
    for row in sheet_credentials.iter_rows(min_row=2, values_only=True):
        user_id, user_name, password = row[0], row[1], row[2]

        # Find username and password fields and login button
        username_field = driver.find_element(By.ID, "user-name")
        password_field = driver.find_element(By.ID, "password")
        login_button = driver.find_element(By.ID, "login-button")

        # Enter credentials and login
        username_field.clear()
        username_field.send_keys(user_name)
        password_field.clear()
        password_field.send_keys(password)
        login_button.click()

        # Check if error message is displayed
        error_message = "N/A"
        try:
            error_message = driver.find_element(By.CSS_SELECTOR, "h3[data-test='error']").text
        except NoSuchElementException:
            pass

        # Write login message to Excel sheet
        sheet_login.append([user_id, error_message])

        # Navigate back to login page
        driver.get("https://www.saucedemo.com/v1/")

    # Save the workbook
    wb.save(credentials_file)

    # Close the browser
    driver.quit()

if __name__ == "__main__":
    CREDENTIALS_FILE = "user_credentials.xlsx"
    LOGIN_SHEET_NAME = "Login Messages"

    login_with_users(CREDENTIALS_FILE, LOGIN_SHEET_NAME)
