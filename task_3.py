"""
Module for automating the login process into a website and scraping product details,
 then saving the scraped data into an Excel file using Selenium and Openpyxl libraries.
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

def login_and_scrape_product_details(url, username, password, output_file):  #pylint:disable=R0914
    """
    This function logs into the given URL, scrapes product details, 
    and saves the data to an Excel file.
   
    Args:
        url (str): The URL to log into.
        username (str): The username to use for login.
        password (str): The password to use for login.
        output_file (str): The path to the Excel file where the product details will be saved.
    
    Returns:
        None
    """
    # Initialize Chrome webdriver
    driver = webdriver.Chrome()
    driver.get(url)

    # Wait for the page to load
    driver.implicitly_wait(10)  # Adjust the waiting time as necessary

    # Find the username and password fields and login button
    username_field = driver.find_element(By.ID, "user-name")
    password_field = driver.find_element(By.ID, "password")
    login_button = driver.find_element(By.ID, "login-button")

    # Enter credentials and login
    username_field.send_keys(username)
    password_field.send_keys(password)
    login_button.click()

    # Wait for the inventory to load
    driver.implicitly_wait(10)

    # Load the existing workbook
    wb = openpyxl.load_workbook(output_file)

    # Check if "Product Details" sheet already exists
    if "Product Details" in wb.sheetnames:
        # If it exists, get the sheet and clear existing content
        sheet_product_details = wb["Product Details"]
        sheet_product_details.delete_rows(2, sheet_product_details.max_row)
    else:
        # If it doesn't exist, create a new sheet
        sheet_product_details = wb.create_sheet("Product Details")
        sheet_product_details['A1'] = "Product ID"
        sheet_product_details['B1'] = "Product Name"
        sheet_product_details['C1'] = "Description"
        sheet_product_details['D1'] = "Price($)"

    # Find all the products
    products = driver.find_elements(By.CLASS_NAME, "inventory_item")

    # Write product details to the sheet
    for idx, product in enumerate(products, start=2):
        product_id = product.find_element(By.CLASS_NAME, "inventory_item_img").find_element(By.TAG_NAME, "a").get_attribute("href").split('=')[-1] #pylint:disable=C0301
        product_name = product.find_element(By.CLASS_NAME, "inventory_item_name").text
        description = product.find_element(By.CLASS_NAME, "inventory_item_desc").text
        price = product.find_element(By.CLASS_NAME, "inventory_item_price").text.replace("$", "")

        sheet_product_details[f'A{idx}'] = product_id
        sheet_product_details[f'B{idx}'] = product_name
        sheet_product_details[f'C{idx}'] = description
        sheet_product_details[f'D{idx}'] = price

    # Save the workbook
    wb.save(output_file)

    # Close the browser
    driver.quit()

if __name__ == "__main__":
    URL = "https://www.saucedemo.com/v1/"
    USERNAME = "standard_user"
    PASSWORD = "secret_sauce"
    OUTPUT_FILE = "user_credentials.xlsx"

    login_and_scrape_product_details(URL, USERNAME, PASSWORD, OUTPUT_FILE)
